#! /usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
#from openpyx.utils import coordinate_from_string
import datetime

def combine():
    wb = load_workbook('courses.xlsx')
    w1 = wb['students']
    w2 = wb['time']
    w3 = wb.copy_worksheet(w1)
    w3.title = 'combine'

    #print(w3)
    course_names = []
    for row in w3['B']:
        #print(row.value)
        #break
        course_names.append(row.value)
    course_names.pop(0)
    #print(course_names)
    w3['D1'].value = w2['C1'].value
    for row in w2.iter_rows(min_row = 2,min_col = 2, max_col = 2):
        #print(type(row[0]))
        for cell in row:
            #print(cell.value)
            index = course_names.index(cell.value)
            cell_name3 = 'D' + str(index + 2)
            #print(cell_name3)
            row_num = cell.row
            cell_name2 = 'C' + str(row_num)
            #print(cell_name2)
            w3[cell_name3].value = w2[cell_name2].value
    wb.save('courses.xlsx')

def row_to_list(ws,n):
    for row in ws.iter_rows(n):
        yield [cell.value for cell in row]

def split():
    wb = load_workbook('courses.xlsx')
    w3 = wb['combine']
    yearset = set()
    for row in w3.iter_rows(min_row = 2, min_col = 1, max_col = 1):
        for cell in row:
            #row_num = cell.row
            #cell_name3 = 'E' + str(row_num)
            #w3[cell_name3].value = cell.value.year
            yearset.add(cell.value.year)
    #wb.save('test.xlsx')

    for year in yearset:
        wb.create_sheet(str(year))
    #wb.save('test.xlsx')

    for row in w3.iter_rows(min_row = 2, min_col = 1, max_col = 1):
        #print(row)
        for cell in row:
            for year in yearset:
                if cell.value.year == year:
                    row_num = cell.row
                    ws_temp = wb.get_sheet_by_name(str(year))
                    #print(ws_temp)
                    n = 'A' + str(cell.row) + ':' + ('D' + str(cell.row))
                    list_to_append = list(row_to_list(w3,n))
                    for items in list_to_append:
                        ws_temp.append(items)

    for year in yearset:
        name = str(year) + '.xlsx'
        wb.save(name)
        wb_temp = load_workbook(name)
        sheets = wb_temp.sheetnames

        for s in sheets:
            if s != str(year):
                sheet_name = wb_temp.get_sheet_by_name(s)
                wb_temp.remove_sheet(sheet_name)
        wb_temp.save(name)

if __name__ == '__main__':
    combine()
    split()
